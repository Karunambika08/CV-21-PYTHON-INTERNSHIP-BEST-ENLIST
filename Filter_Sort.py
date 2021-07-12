Filter (Where Clause) & Sort¶

1. Create an Excel with data of Student database and add all the values which is required for student management database, Read the excel file and add the datas into a DB (using script)

In [1]:

import openpyxl

In [4]:

path = "students.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row = 5, column = 3)

print(cell_obj.value)

divya

In [9]:

# print one student complete record from excel sheet

for i in range(1,11):

    cell_obj = sheet_obj.cell(row = 5, column = i)

    print(cell_obj.value)

4

104

divya

9.8

9.5

8.8

9.366666667

9876434321

divya@gmail.com

saravanan

In [10]:

import mysql.connector

mydb = mysql.connector.connect(

  host="localhost",

 user="root",

  password="1234",

)

mycursor = mydb.cursor()

print(mydb)

<mysql.connector.connection_cext.CMySQLConnection object at 0x000001ECE4D6E580>

In [12]:

dbse = mydb.cursor()

dbse.execute("CREATE DATABASE Students_Management_System")

In [13]:

dbse = mydb.cursor()

dbse.execute("SHOW DATABASES")

for entry in dbse:

  print(entry)

('doctor',)

('doctors1',)

('grocerystore',)

('information_schema',)

('mydatabase',)

('mysql',)

('performance_schema',)

('sakila',)

('students_management_system',)

('sys',)

('world',)

In [71]:

mydb = mysql.connector.connect(

  host="localhost",

 user="root",

  password="1234",

  database="students_management_system"

)

dbse = mydb.cursor()

dbse.execute("CREATE TABLE student3 (roll_no INT(10),Reg_no INT(10),NAME VARCHAR(255), semester1 INT(25),semester2 INT(25),semester3 INT(25), CGPA INT(35) ,PHONE_NUMBER INT,email_id VARCHAR(255))")

In [72]:

dbse = mydb.cursor()

dbse.execute("SHOW TABLES")

for value in dbse:

  print(value)

('student1',)

('student2',)

('student3',)

('students',)

In [73]:

cur = mydb.cursor()

cur.execute('SELECT * FROM student3')

for row in cur:

    print(row)

In [74]:

import pandas as pd

df = pd.read_excel('students.xlsx')

In [75]:

import xlrd

import MySQLdb

In [76]:

xl_sheet = xlrd.open_workbook("students.xlsx")

xl_sheet

Out[76]:

<xlrd.book.Book at 0x1ece7ca2430>

In [77]:

sheet_name =xl_sheet.sheet_names()

sheet_name

Out[77]:

['students']

In [78]:

In [92]:

mydb = mysql.connector.connect(

  host="localhost",

 user="root",

  password="1234",

  database="students_management_system"

)

cur = mydb.cursor()

for s in range(0,1):

    sheet=xl_sheet.sheet_by_index(s)

    sql= "INSERT INTO student3(roll_no,Reg_no,NAME,semester1,semester2 ,semester3 , CGPA ,email_id) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)"

    for r in range(1,sheet.nrows):

        roll_no =sheet.cell(r,0).value

        Reg_no =sheet.cell(r,1).value

        NAME =sheet.cell(r,2).value

        semester1 =sheet.cell(r,3).value

        semester2 =sheet.cell(r,4).value

        semester3 =sheet.cell(r,5).value

        CGPA =sheet.cell(r,6).value

        email_id=sheet.cell(r,7).value

        values =(roll_no ,Reg_no,NAME ,semester1,semester2 ,semester3 , CGPA ,email_id)

        

        cur.execute(sql,values)

mydb.commit()

In [93]:

mycursor = mydb.cursor()

mycursor.execute("SELECT * FROM student3")

myresult = mycursor.fetchall()

for x in myresult:

    print(x)

(1, 101, 'Anu', 9, 7, 7, 8, None, 'qwer@gmail.com')

(1, 101, 'Anu', 9, 7, 7, 8, None, 'qwer@gmail.com')

(1, 101, 'Anu', 9, 7, 7, 8, None, 'qwer@gmail.com')

(1, 101, 'Anu', 9, 7, 7, 8, None, 'qwer@gmail.com')

(1, 101, 'Anu', 9, 7, 7, 8, None, 'qwer@gmail.com')

(1, 101, 'Anu', 9, 7, 7, 8, None, 'qwer@gmail.com')

(2, 102, 'bargavi', 9, 9, 9, 9, None, 'bargavi@gmail.com')

(3, 103, 'chitra', 7, 9, 9, 8, None, 'shitra@gmail.com')

(4, 104, 'divya', 10, 10, 9, 9, None, 'divya@gmail.com')

(5, 105, 'elikya', 9, 6, 10, 8, None, 'elikya@gmail.com')

(6, 106, 'fathima', 7, 9, 9, 8, None, 'fathimababu@gmail.com')

(7, 107, 'geeta', 8, 8, 9, 8, None, 'geetasaea@gmail.com')

(8, 108, 'harshita', 9, 8, 9, 8, None, 'harshittta@gmail.com')

(9, 109, 'iswarya', 8, 8, 9, 8, None, 'iswarya@gmail.com')

(10, 110, 'jeevika', 9, 9, 8, 8, None, 'jeevikamanvi@gmail.com')

(11, 111, 'koushilya', 8, 9, 8, 8, None, 'kuoshi@gmail.com')

(12, 112, 'lavanya', 3, 8, 10, 7, None, 'lavanya@gmail.com')

(13, 113, 'malavika', 6, 8, 9, 8, None, 'malavika@gmail.com')

(14, 114, 'naveena', 8, 7, 7, 7, None, 'naveena@gmail.com')

(15, 115, 'priya', 5, 8, 7, 7, None, 'priya@gmail.com')

(1, 101, 'Anu', 9, 7, 7, 8, None, 'qwer@gmail.com')

(2, 102, 'bargavi', 9, 9, 9, 9, None, 'bargavi@gmail.com')

(3, 103, 'chitra', 7, 9, 9, 8, None, 'shitra@gmail.com')

(4, 104, 'divya', 10, 10, 9, 9, None, 'divya@gmail.com')

(5, 105, 'elikya', 9, 6, 10, 8, None, 'elikya@gmail.com')

(6, 106, 'fathima', 7, 9, 9, 8, None, 'fathimababu@gmail.com')

(7, 107, 'geeta', 8, 8, 9, 8, None, 'geetasaea@gmail.com')

(8, 108, 'harshita', 9, 8, 9, 8, None, 'harshittta@gmail.com')

(9, 109, 'iswarya', 8, 8, 9, 8, None, 'iswarya@gmail.com')

(10, 110, 'jeevika', 9, 9, 8, 8, None, 'jeevikamanvi@gmail.com')

(11, 111, 'koushilya', 8, 9, 8, 8, None, 'kuoshi@gmail.com')

(12, 112, 'lavanya', 3, 8, 10, 7, None, 'lavanya@gmail.com')

(13, 113, 'malavika', 6, 8, 9, 8, None, 'malavika@gmail.com')

(14, 114, 'naveena', 8, 7, 7, 7, None, 'naveena@gmail.com')

(15, 115, 'priya', 5, 8, 7, 7, None, 'priya@gmail.com')

In [94]:

mycursor = mydb.cursor()

mycursor.execute("SELECT NAME FROM student3 WHERE CGPA >6")

myresult = mycursor.fetchall()

for x in myresult:

    print(x)

('Anu',)

('Anu',)

('Anu',)

('Anu',)

('Anu',)

('Anu',)

('bargavi',)

('chitra',)

('divya',)

('elikya',)

('fathima',)

('geeta',)

('harshita',)

('iswarya',)

('jeevika',)

('koushilya',)

('lavanya',)

('malavika',)

('naveena',)

('priya',)

('Anu',)

('bargavi',)

('chitra',)

('divya',)

('elikya',)

('fathima',)

('geeta',)

('harshita',)

('iswarya',)

('jeevika',)

('koushilya',)

('lavanya',)

('malavika',)

('naveena',)

('priya',)

In [103]:

mydb.commit()

 # Close the database connection

mydb.close()

In [ ]:
